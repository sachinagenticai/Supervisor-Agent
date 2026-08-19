from __future__ import annotations

import json
from datetime import datetime, timezone
from os import replace
from pathlib import Path
from threading import RLock
from typing import Any, Callable

from filelock import FileLock, Timeout
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


EXCEL_SCHEMA_VERSION = "3.1"

EXCEL_HEADERS: dict[str, list[str]] = {
    "_meta": ["key", "value", "updated_at"],
    "application_user": [
        "id", "google_subject_id", "email", "display_name", "profile_image_url",
        "created_at", "last_login_at",
    ],
    "agent_registry": [
        "id", "agent_code", "agent_name", "description", "version", "owner", "lifecycle_status",
        "capabilities", "source_systems", "record_types", "routing_key_hints", "rule_pack_id",
        "tool_code", "plugin", "judge_rubric", "success_tag", "thresholds", "required_evidence",
        "escalation_policy", "enabled", "created_at", "updated_at",
    ],
    "validation_record": [
        "id", "external_reference", "source_system", "record_type", "record_title",
        "expected_agent_code", "payload", "metadata", "active", "created_at",
    ],
    "validation_run": [
        "id", "record_id", "initiated_by_user_id", "comments", "execution_status",
        "detected_agent_code", "selected_tool_code", "routing_reason", "routing_confidence",
        "routing_method", "routing_candidates", "final_verdict", "business_decision",
        "final_reason", "final_tag", "final_confidence", "assurance_band", "recommended_action",
        "data_completeness", "score_breakdown", "disagreement_detected", "degraded_mode",
        "context_snapshot", "memory_snapshot", "governance", "remediation", "started_at",
        "completed_at", "error_message",
    ],
    "rule_result": [
        "id", "run_id", "rule_code", "rule_name", "severity", "passed", "mandatory",
        "evidence", "message", "tag", "created_at",
    ],
    "llm_judgement": [
        "id", "run_id", "model_name", "judge_verdict", "confidence", "reason", "analysis",
        "findings", "recommendations", "quality_dimensions", "focus_area_addressed",
        "degraded_mode", "raw_response", "prompt_version", "created_at",
    ],
    "audit_event": ["id", "run_id", "user_id", "event_type", "event_details", "created_at"],
    "connector_sync": [
        "id", "connector_code", "sync_status", "records_read", "records_written", "details",
        "started_at", "completed_at",
    ],
}

_process_lock = RLock()


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def json_dumps(value: Any) -> str:
    return json.dumps(value, default=str, ensure_ascii=False, separators=(",", ":"))


def json_loads(value: Any, default: Any | None = None) -> Any:
    fallback = {} if default is None else default
    if value in (None, ""):
        return fallback
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(str(value))
    except (json.JSONDecodeError, TypeError):
        return fallback


def _atomic_save_workbook(workbook: Workbook, target_path: Path) -> None:
    target_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = target_path.with_name(f"{target_path.stem}.{datetime.now().timestamp():.0f}.tmp{target_path.suffix}")
    workbook.save(temp_path)
    replace(temp_path, target_path)


def _style_sheet(ws: Worksheet, headers: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for index, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=index)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = min(max(len(header) + 3, 14), 32)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=max(ws.max_row, 1), column=len(headers)).coordinate}"


def _migrate_sheet(wb: Workbook, sheet_name: str, headers: list[str]) -> None:
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        _style_sheet(ws, headers)
        return

    ws = wb[sheet_name]
    old_headers = [ws.cell(row=1, column=index).value for index in range(1, ws.max_column + 1)]
    old_headers = [str(value) if value is not None else "" for value in old_headers]
    if old_headers == headers:
        _style_sheet(ws, headers)
        return

    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in values):
            continue
        rows.append({old_headers[index]: value for index, value in enumerate(values) if index < len(old_headers)})

    position = wb.sheetnames.index(sheet_name)
    wb.remove(ws)
    new_ws = wb.create_sheet(sheet_name, position)
    new_ws.append(headers)
    for row in rows:
        new_ws.append([row.get(header) for header in headers])
    _style_sheet(new_ws, headers)


def initialize_excel_workbook(path: str | Path, reset: bool = False) -> None:
    workbook_path = Path(path)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    with _process_lock:
        if workbook_path.exists() and not reset:
            wb = load_workbook(workbook_path)
        else:
            wb = Workbook()
            wb.remove(wb.active)
        for sheet_name, headers in EXCEL_HEADERS.items():
            _migrate_sheet(wb, sheet_name, headers)
        meta = wb["_meta"]
        existing = {
            str(meta.cell(row=row, column=1).value): row
            for row in range(2, meta.max_row + 1)
            if meta.cell(row=row, column=1).value
        }
        for key, value in {
            "schema_version": EXCEL_SCHEMA_VERSION,
            "storage_mode": "excel_single_instance",
            "description": "Enterprise AI Supervisor Excel-backed controlled deployment store",
        }.items():
            if key in existing:
                row = existing[key]
                meta.cell(row=row, column=2).value = value
                meta.cell(row=row, column=3).value = now_iso()
            else:
                meta.append([key, value, now_iso()])
        _atomic_save_workbook(wb, workbook_path)
        wb.close()


class ExcelDataStore:
    def __init__(self, path: str | Path):
        self.path = Path(path)
        initialize_excel_workbook(self.path)
        self.workbook = load_workbook(self.path)
        self.dirty = False

    def save(self) -> None:
        if not self.dirty:
            return
        self.upsert(
            "_meta", "key", "last_saved_at",
            {"key": "last_saved_at", "value": now_iso(), "updated_at": now_iso()},
        )
        _atomic_save_workbook(self.workbook, self.path)
        self.dirty = False

    def close(self) -> None:
        self.workbook.close()

    def sheet(self, name: str) -> Worksheet:
        if name not in EXCEL_HEADERS:
            raise ValueError(f"Unknown Excel sheet: {name}")
        return self.workbook[name]

    def headers(self, sheet_name: str) -> list[str]:
        return EXCEL_HEADERS[sheet_name]

    def rows(self, sheet_name: str) -> list[dict[str, Any]]:
        ws = self.sheet(sheet_name)
        headers = self.headers(sheet_name)
        records: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, max_col=len(headers), values_only=True):
            if not any(cell not in (None, "") for cell in row):
                continue
            records.append({header: row[index] for index, header in enumerate(headers)})
        return records

    def insert(self, sheet_name: str, row: dict[str, Any]) -> None:
        ws = self.sheet(sheet_name)
        headers = self.headers(sheet_name)
        ws.append([self._normalize_cell(row.get(header)) for header in headers])
        self.dirty = True

    def upsert(self, sheet_name: str, key: str, key_value: Any, values: dict[str, Any]) -> dict[str, Any]:
        ws = self.sheet(sheet_name)
        headers = self.headers(sheet_name)
        key_col = headers.index(key) + 1
        for row_index in range(2, ws.max_row + 1):
            if ws.cell(row=row_index, column=key_col).value == key_value:
                for field, value in values.items():
                    if field in headers:
                        ws.cell(row=row_index, column=headers.index(field) + 1).value = self._normalize_cell(value)
                self.dirty = True
                return self.find_one(sheet_name, lambda row: row.get(key) == key_value) or values
        self.insert(sheet_name, values)
        return values

    def update(self, sheet_name: str, key: str, key_value: Any, values: dict[str, Any]) -> None:
        ws = self.sheet(sheet_name)
        headers = self.headers(sheet_name)
        key_col = headers.index(key) + 1
        for row_index in range(2, ws.max_row + 1):
            if ws.cell(row=row_index, column=key_col).value == key_value:
                for field, value in values.items():
                    if field in headers:
                        ws.cell(row=row_index, column=headers.index(field) + 1).value = self._normalize_cell(value)
                self.dirty = True
                return
        raise ValueError(f"No row found in {sheet_name} where {key}={key_value}")

    def find_one(self, sheet_name: str, predicate: Callable[[dict[str, Any]], bool]) -> dict[str, Any] | None:
        return next((row for row in self.rows(sheet_name) if predicate(row)), None)

    def delete_all(self, sheet_name: str) -> None:
        ws = self.sheet(sheet_name)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
            self.dirty = True

    @staticmethod
    def _normalize_cell(value: Any) -> Any:
        if isinstance(value, (dict, list, tuple, set)):
            return json_dumps(value)
        if isinstance(value, datetime):
            return value.isoformat(timespec="seconds")
        return value


class ExcelTransaction:
    def __init__(self, path: str | Path, timeout_seconds: int = 30):
        self.path = Path(path)
        self.timeout_seconds = timeout_seconds
        self.store: ExcelDataStore | None = None
        self.file_lock = FileLock(str(self.path) + ".lock", timeout=timeout_seconds)

    def __enter__(self) -> ExcelDataStore:
        _process_lock.acquire()
        try:
            self.file_lock.acquire()
            self.store = ExcelDataStore(self.path)
            return self.store
        except Timeout as exc:
            _process_lock.release()
            raise TimeoutError(
                f"Excel store is busy after waiting {self.timeout_seconds} seconds. Retry the operation."
            ) from exc
        except Exception:
            if self.file_lock.is_locked:
                self.file_lock.release()
            _process_lock.release()
            raise

    def __exit__(self, exc_type: Any, exc: Any, tb: Any) -> None:
        try:
            if self.store is not None and exc_type is None:
                self.store.save()
        finally:
            if self.store is not None:
                self.store.close()
            if self.file_lock.is_locked:
                self.file_lock.release()
            _process_lock.release()
